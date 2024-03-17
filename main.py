import json
import os
import time
import pytz
import sqlite3
import datetime
import openpyxl
import urllib.request
from pytz import utc
from flask import Flask, render_template,request
from linebot import LineBotApi, WebhookHandler

app = Flask(__name__)

CHANNEL_ACCESS_TOKEN = "3rWORuikyamxptl2guKgn4kIyCYVFiYOZjfRSZ3hgnTKl5kIx3jhvkq6FiLZlOo6HJS2g3WFQPdGTeA2hIDiNsR8jwdyKcC95QwKISrGcu/nZxCINEcE8goEWHXk6c/frIE56ge/OGbewI9uzCwOGAdB04t89/1O/w1cDnyilFU="
Channel_secret = "60c382a7a969d7c0a80e26d8c792d34f"
SPREADSHEET_ID = "1Xx0TnoamQkH0_gaLOOzMyAmGBUwufYKGURHVxrzfOiw"

# 建立 LINE Bot API 物件
line_bot_api = LineBotApi(CHANNEL_ACCESS_TOKEN)

# 建立 WebhookHandler 物件
handler = WebhookHandler(Channel_secret)

# 設定要忽略的字詞
IGNORE_WORDS = ["help", "?"]

# 設定停止提醒的關鍵字
STOP_ALARM_WORD = "ok"

# 取得試算表的資料
# 建立 Workbook 物件
workbook = openpyxl.load_workbook("C:/Users/Alan6/OneDrive/桌面/data.xltm")
# 取得工作表
sheet = workbook.active
# 讀取 A1 到 B10 的資料
data = list(sheet.values)[0:10]
# 輸出資料
for row in data:
    print(row)

def update_sheet_data(spreadsheet_id, data):
  """
  使用 Google Sheets API 更新試算表
  """
  # 建立 Google Sheets API 服務物件
  service = build('sheets', 'v4', credentials=YOUR_CREDENTIALS)
  # 更新試算表資料
  service.spreadsheets().values().update(spreadsheetId=spreadsheet_id, range='A1:B10', body={'values': data}).execute()
  # 設定試算表 ID
  SPREADSHEET_ID = "1Xx0TnoamQkH0_gaLOOzMyAmGBUwufYKGURHVxrzfOiw"
  # 取得試算表的資料
  data = get_sheet_data(SPREADSHEET_ID)
  # 更新試算表的資料 
  update_sheet_data(SPREADSHEET_ID, data)

def handle_user_message(event):
  """
  處理使用者傳送的訊息
  """
  # 取得使用者 ID 和傳送的訊息
  user_id = event["source"]["userId"]
  message_text = event["message"]["text"]
  # 忽略指定的字詞
  if message_text.lower() in IGNORE_WORDS:
    return
  # 處理停止提醒的指令
  if message_text.lower() == STOP_ALARM_WORD.lower():
    stop_alarm()
    return
  # 使用 sqlite3 存取使用者回答
  connection = sqlite3.connect("your_database.sqlite")
  cursor = connection.cursor()
  # 取得使用者目前回答到第幾題
  answer_index = cursor.execute("SELECT answer_index FROM users WHERE user_id = ?", (user_id,)).fetchone()[0]
  # 設定使用者回答的答案
  cursor.execute("UPDATE users SET answer = ? WHERE user_id = ?", (message_text, user_id,))
  connection.commit()
  # 如果使用者已經回答完所有問題，就設定提醒時間
  if next_question_index == 1:
    # 使用 datetime 設定提醒時間
    now = datetime.datetime.now(utc)
    alarm_time = now + datetime.timedelta(hours=1)
    # 使用 requests 發送提醒
    headers = {
      "Authorization": "Bearer <YOUR_ACCESS_TOKEN>"
    }
    data = {
      "to": user_id,
      "messages": [
        {
          "type": "text",
          "text": "請記得回答問題！"
        }
      ]
    }
    response = requests.post("https://api.line.me/v2/bot/message/push", headers=headers, data=json.dumps(data))
    return
  # 傳送下一題給使用者
  send_question(user_id, next_question_index)

def get_user_answer(user_id, message_text):
  """
  取得使用者回答到第幾題
  """
  # 使用 openpyxl 讀取 Excel 檔案
  workbook = openpyxl.load_workbook("your_file.xlsx")
  sheet = workbook.active
  # 尋找使用者資料
  for i in range(2, len(sheet.values)):
    if sheet.values[i][0] == user_id and sheet.values[i][len(sheet.values[0]) - 1] == "":
      # 尋找使用者尚未回答的題目
      for j in range(1, len(sheet.values[0]) - 1):
        if sheet.values[i][j] == "":
          # 更新使用者回答
          sheet.values[i][j] = message_text
          # 如果使用者已經回答完所有問題，就設定提醒時間
          if j + 3 == len(sheet.values[0]):
            return i, 0
          # 否則傳回下一題的索引
          else:
            return i, j + 2
  # 使用者資料不存在或已回答完所有問題
  return None, 1

def set_alarm(answer_index):
  """
  設定提醒時間
  """
  # 使用 openpyxl 讀取 Excel 檔案
  workbook = openpyxl.load_workbook("your_file.xlsx")
  sheet = workbook.active
  # 取得提醒時間
  alarm_time_str = sheet.values[answer_index][1] + " " + sheet.values[answer_index][2]
  # 使用 pytz 解析使用者輸入的時間
  alarm_time = datetime.datetime.strptime(alarm_time_str, "%Y/%m/%d %H:%M").replace(tzinfo=utc)
  # 使用者的提醒時間如果已經過了，就跳過設定提醒的步驟
  if alarm_time < datetime.datetime.now(utc):
    return
  # 設定提醒
  # ...
  workbook.save("your_file.xlsx")

def stop_alarm():
  """
  停止提醒
  """
  # 取得試算表資料
  sheet_data = get_sheet_data()
  # 尋找尚未提醒的使用者
  for i in range(2, len(sheet_data)):
    if sheet_data[i][4] == 0:
      # 檢查提醒時間是否已經過了
      alarm_time = datetime.datetime.strptime(sheet_data[i][1] + " " + sheet_data[i][2], "%Y/%m/%d %H:%M")
      if alarm_time < datetime.datetime.now():
        # 取消提醒
        # ...
        # 更新提醒狀態
        sheet_data[i][4] = 1
  # 儲存試算表資料
  save_sheet_data(sheet_data)

def send_question(user_id, question_index):
  """
  傳送問題給使用者
  """
  # 取得試算表資料
  sheet_data = get_sheet_data()
  # 取得問題
  question = sheet_data[0][question_index - 1]
  # 根據問題類型傳送不同的訊息
  if question_index == 2:
    # 日期問題
    send_date_picker(user_id, question)
  elif question_index == 3:
    # 時間問題
    send_time_picker(user_id, question)
  else:
    # 其他問題
    send_confirm_message(user_id, question)

def send_date_picker(user_id, question):
  """
  傳送日期選擇器給使用者
  """
  # 設定日期選擇器的參數
  data = {
    "type": "datetimepicker",
    "label": question,
    "data": "DateMessage",
    "mode": "date",
    "id": "date_picker",
  }
  # 傳送日期選擇器
  send_message(user_id, [data])

# 傳送確認訊息
def send_confirm_message(user_id, question):
    # 設定確認訊息的參數
    actions = {
    {
        "type": "postback",
        "label": "確認",
        "data": "DecideConfirm",
        "text": "確認",
    },
    {
        "type": "postback",
        "label": "取消",
        "data": "DecideCancel",
        "text": "取消",
    },
    # 傳送確認訊息
    send_message(user_id, [data])}

def send_message(user_id, messages):
  """
  傳送訊息給使用者
  """
  # 設定傳送訊息的參數
  headers = {
    "Authorization": "Bearer " + CHANNEL_ACCESS_TOKEN,
    "Content-Type": "application/json",
  }
  body = {
    "to": user_id,
    "messages": messages,
  }
  # 傳送訊息
  urllib.request.urlopen(
    "https://api.line.me/v2/bot/message/push",
    data=json.dumps(body).encode("utf-8"),
    headers=headers,
  )

@app.route('/callback', methods=['POST'])
def callback(request):
    # 驗證請求
    if request.headers['X-Line-Signature'] != 'Channel_secret':
        return '403 Forbidden', 403

    # 解析請求
    events = request.json['events']

    # 處理事件
    for event in events:
        if event['type'] == 'message':
            # 回覆訊息
            reply_message = {
                'type': 'text',
                'text': event['message']['text']
            }
            client.reply_message(event['replyToken'], reply_message)

    return '200 OK', 200
                      
if __name__ == "__main__":
      app.run()